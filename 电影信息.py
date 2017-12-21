# -*- coding: utf-8 -*-
"""
Created on Wed Nov 15 13:16:02 2017

@author: Administrator
"""

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import re
import xlwt
import numpy as np
import threading
from openpyxl.workbook import Workbook
from bs4 import BeautifulSoup as Soup
import pandas

pathes = 'http://58921.com/boxoffice/wangpiao/20160101'
driver = webdriver.Firefox()
driver.get('http://58921.com/boxoffice/wangpiao/20160101')
source = driver.page_source


def filminfoperday():
    big_month = (1,3,5,7,8,10,12)
    film_all = []
    for i in range(12):
        if i+1 in big_month:
            ran = 31
        elif i+1 == 2:
            ran = 28
        else:ran=30
        if i<9:
            stri = '0' + str(i+1)
        else: stri = str(i+1)
        for j in range(ran):
            if j<9:
                strj = '0' + str(j+1)
            else:strj = str(j+1)
            info_day = 'http://58921.com/boxoffice/wangpiao/2016' + stri + strj
            driver.get(info_day)
            source = driver.page_source
            film_dailyindex = re.findall(r'<a href="/content/film/\d+/boxoffice"',source)
            film_all.extend(film_dailyindex)
            time.sleep(.1)
    film_all = list(map(lambda x: r'http://58921.com' + x.split(r'"')[-2],film_all))
    allfilmurl = list({}.fromkeys(film_all).keys() )    
    return allfilmurl

allfilm = filminfoperday()
'''
name = re.findall(r'<a href="/film/\d+/boxoffice" title=".+?">',source)
for i in range(len(name)):
    name[i] = name[i].split(r'"')[-2]

totalbox = re.findall(r'<a href="/content/film/\d+/boxoffice" title="总票房明细">.+?<',source)
for i in range(len(totalbox)):
    totalbox[i] = totalbox[i].split(r'>')[-1].split(r'<')[-2]
    if totalbox[i][-1] == u'亿':
        totalbox[i] = float(totalbox[i][:-1])*1e8
    else:totalbox[i] = float(totalbox[i][:-1])*1e4
'''







def down_table(url):
    driver.get(url)
    source = driver.page_source
    soup = Soup(source,'lxml')
    table = Soup(str(soup.table()),'lxml')

    table_list = list(table.find_all('td'))
    final_list = list(map(lambda x: str(x).split(r'<')[int((len(str(x).split(r'<'))+1)/2)-1].split(r'>')[-1],table_list))
    
    col_even1 = soup.find_all('tr',class_ ="even")[0]
    judge_a = Soup(str(col_even1),'lxml')   
    judge_list = list(judge_a.find_all('td'))
    judge_list = list(map(lambda x: str(x).split(r'<')[int((len(str(x).split(r'<'))+1)/2)-1].split(r'>')[-1],judge_list))     
    if len(judge_list) != 19:
        pad  = ['--']
        dif = 19 - len(judge_list)
        final_list[19+len(judge_list):19+len(judge_list)] = pad*dif
    name = re.findall(r'<h2>.+?</h2>',source)[0].split(r'<')[-2].split(r'>')[1]
    book=Workbook()
    sheet1=book.active
    sheet1.title = "电影信息"
    sheet1.merge_cells('A1:S1')
    sheet1.cell(row = 1,column=1 ,value = name) 
    head = ['时间','网票','哈票','万达','金逸','淘电影','星美']
    for i in range(len(head)):
        if i == 0:
            sheet1.cell(row = 2,column=i+1,value = head[i]) 
        else:
            sheet1.cell(row = 2,column=3*i-1,value = head[i]) 
    sheet1.merge_cells('B2:D2')
    sheet1.merge_cells('E2:G2')
    sheet1.merge_cells('H2:J2')
    sheet1.merge_cells('K2:M2')
    sheet1.merge_cells('N2:P2')
    sheet1.merge_cells('Q2:S2')
    for i in range(int(len(final_list)/19)):
        for j in range(19):
            sheet1.cell(row = i+3,column=j+1 ,value = final_list[19*i+j]) 
    book.save(name + '.xlsx')  
    splitcri = u'每日票房数据统计'
    return name.split(splitcri)[0]    

names = list(map(down_table,allfilm))
splitcri = u'每日票房数据统计'
names = list(map(lambda x: x.split(splitcri)[0],names))   
    


    






