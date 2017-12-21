# -*- coding: utf-8 -*-
"""
Created on Thu Nov  9 20:19:58 2017

@author: Administrator
"""

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import re
import xlwt
import numpy as np
import threading
from openpyxl import Workbook

'''
模拟登陆
'''
driver = webdriver.Firefox()
driver.get('https://www.douban.com/accounts/login?source=movie')
time.sleep(1)
driver.find_element_by_id('email').send_keys('*********')
time.sleep(1)
driver.find_element_by_id('password').send_keys('*********')
time.sleep(1)
driver.find_element_by_name('login').click()
time.sleep(1)
driver.find_element_by_id('email').clear()
time.sleep(1)
driver.find_element_by_id('email').send_keys('********')
time.sleep(1)
driver.find_element_by_id('password').send_keys('********')
time.sleep(1)
driver.find_element_by_name('login').click()
cookies = driver.get_cookies()

'''
增加线程
'''
driver1 = webdriver.Firefox()
driver1.get('https://www.douban.com/accounts/login?source=movie')
time.sleep(1)
driver1.find_element_by_id('email').send_keys('*********')
time.sleep(1)
driver1.find_element_by_id('password').send_keys('*********')
time.sleep(1)
driver1.find_element_by_name('login').click()
time.sleep(1)
driver1.find_element_by_id('email').clear()
time.sleep(1)
driver1.find_element_by_id('email').send_keys('********')
time.sleep(1)
driver1.find_element_by_id('password').send_keys('********')
time.sleep(1)
driver1.find_element_by_name('login').click()
cookies = driver.get_cookies()

'''
增加线程
下为弃用线程，线程过多会导致豆瓣判断为机器人，封IP
driver2 = webdriver.Firefox()




driver4 = webdriver.Firefox()
driver4.get('https://www.douban.com/accounts/login?source=movie')
time.sleep(1)
driver4.find_element_by_id('email').send_keys('********')
time.sleep(1)
driver4.find_element_by_id('password').send_keys('********')
time.sleep(1)
captcha_id = input('验证码为：')
driver4.find_element_by_id('captcha_field').send_keys(captcha_id)
time.sleep(1)
driver4.find_element_by_name('login').click()
time.sleep(1)
driver4.find_element_by_id('email').clear()
time.sleep(1)
driver4.find_element_by_id('email').send_keys('********')
time.sleep(1)
driver4.find_element_by_id('password').send_keys('********')
time.sleep(1)
captcha_id = input('验证码为：')
driver4.find_element_by_id('captcha_field').send_keys(captcha_id)
time.sleep(1)
driver4.find_element_by_name('login').click()
cookies = driver.get_cookies()
'''
'''
找到电影条目
'''
pathes = []
for i in range(21):
    url = 'https://movie.douban.com/annual2016/?source=navigation#'+str(i+1)
    driver.get(url)
    source = driver.page_source
    path = re.findall("https://movie.douban.com/subject/\d+",source)
    pathes.extend(path)
    time.sleep(5)

pathes_new1 = []
for i in pathes:
    if not i in pathes_new1:
        pathes_new1.append(i)

pathes_new2 = []
for i in pathes_new1:
    if not i in pathes_new2:
        pathes_new2.append(i)
'''
提取评论
'''
def commentsinfo(url,driver_in):
    driver = eval(driver_in)
    driver.get(url)
    html = driver.page_source
    rank_pattern = re.compile(u'看过</span>\n.+')
    comments_pattern = re.compile(r'<p class="">.+')
    votes_pattern = re.compile(r'<span class="votes">\d+')
    time_pattern = re.compile(r'<span class="comment-time " title="\d{4}-\d{2}-\d{2}')
    rank = re.findall(rank_pattern,html) 
    comments = re.findall(comments_pattern,html)
    votes = re.findall(votes_pattern,html)
    date = re.findall(time_pattern,html)
    for i in range(len(comments)):
        comments[i] = comments[i].split('>')[1]
    for i in range(len(rank)):
        a = 'span class'
        b = rank[i].split(a)[1]
        if b[2:9] !='allstar':
            b='None'
        rank[i] = b
    for i in range(len(rank)):
        if rank[i] != 'None':
            c = 'allstar'
            rank[i] = rank[i].split(c)[1].split(' ')[0]
    for i in range(len(votes)):
        votes[i] = votes[i].split('>')[1]
    for i in range(len(date)):
        date[i] = date[i].split('"')[-1]
    return rank,comments,votes,date
'''
电影信息
'''
def findfilminfo(url,driver_in):
    driver = eval(driver_in)
    driver.get(url)
    html = driver.page_source
    name_pattern = re.compile(r'<span property="v:itemreviewed">.+</span>')
    name = re.findall(name_pattern,html)
    score_pattern = re.compile(r'<strong class="ll rating_num" property="v:average">.{3}')
    score = re.findall(score_pattern,html)
    name = name[0].split('>')[1].split('<')[0]
    score = score[0].split('>')[-1]
    code = url.split(r'/')[-1]
    comment_num = int(re.findall(u'全部.+条',html)[0].split(' ')[1])
    return name,score,code,comment_num

'''
开始抓取电影评论
'''
def commentcaptcha(a,b,driver_in):
    for j in range(a,b):
        name,score,code,comment_num = findfilminfo(pathes_new2[j],driver_in)  
        ranks = []
        comments = []
        votes = []
        dates = []
        time_j = time.time()
        for i in range(0,int(comment_num/20)):
                driver = eval(driver_in)
                source = driver.page_source
                if re.findall(u'你访问豆瓣的方式有点像机器人程序',source):
                    captcha = input(driver_in + '请输入验证码：')
                    driver.find_element_by_name('captcha-solution').send_keys(captcha)
                    driver.find_element_by_xpath(u'//*[@value="我真的不是程序"]').click()
                url = pathes_new2[j] + r'/comments?start=' + str(20*i) + r'&limit=20&sort=new_score&status=P&percent_type='
                rank,comment,vote,date = commentsinfo(url,driver_in)
                ranks.extend(rank)
                comments.extend(comment)
                votes.extend(vote)
                dates.extend(date)
                if i % 50 == 0:
                    time_i = time.time()
                    time_dif = time_i - time_j
                    hour = int(time_dif/3600)
                    minute = int((time_dif - hour*3600)/60)
                    second = int(time_dif - hour*3600 - minute*60)
                    print('----------------------------------------------------------')
                    print('线程: '+ driver_in)
                    print('这是第'+str(j)+'个')
                    print('该电影已用时：'+ str(hour) + '小时' + str(minute) + '分钟' + str(second) + '秒')
                    print('ranks的长度为: ' + str(len(ranks)) + ' comments的长度为: ' + str(len(comments))
                    +' votes的长度为: ' + str(len(votes))+' dates的长度为: ' + str(len(dates)))
                    print('----------------------------------------------------------')
                if len(ranks) != len(comments):
                    print('第 ' + str(i*20) + '页出错了')
                    pad = ['None'] * (len(comments)-len(ranks))
                    ranks.extend(pad)
                time.sleep(1+float(np.random.random(1)))
        book=Workbook()
        sheet1=book.active
        sheet1.title = "评论信息"
        col1=('打分','评论日期','赞同数','评论内容')
        for i in range(0,4):
            sheet1.cell(row = 1,column = i+1 ,value = col1[i])
        for i in range(len(comments)):
            try:
             sheet1.cell(row = i+2,column=1 ,value = ranks[i])
             sheet1.cell(row = i+2,column=2 ,value = dates[i])
             sheet1.cell(row = i+2,column=3 ,value = votes[i])
             sheet1.cell(row = i+2,column=4 ,value = comments[i])
            except:
                print('好像哪里错了')
                continue
        sheet2=book.create_sheet(title="电影信息")
        col2=('电影名','评分','编号','评论数')
        for i in range(0,4):
            sheet2.cell(row =1,column =i+1,value = col2[i])
        sheet2.cell(row =2,column=1 , value = name)
        sheet2.cell(row =2,column=2 , value = score)
        sheet2.cell(row =2,column=3 , value = code)
        sheet2.cell(row =2,column=4 , value = comment_num)
        book.save(code + '.xlsx')
        time_end = time.time()
        time_dif = time_end - time_j
        hour = int(time_dif/3600)
        minute = int((time_dif - hour*3600)/60)
        second = int(time_dif - hour*3600 - minute*60)
        print('该电影总共用时：'+ str(hour) + '小时' + str(minute) + '分钟' + str(second) + '秒')
        time.sleep(5)
'''
下为弃用的 write excel 方式：该方式写入EXCEL只能写入最多65536行，有些爬出的数据超过10W条，无法存入
'''
        #book=xlwt.Workbook(encoding='utf-8',style_compression=0)
        #sheet1=book.add_sheet('sheet1',cell_overwrite_ok=True)
        #col1=('打分','评论日期','赞同数','评论内容')
        #for i in range(0,4):
            #sheet1.write(0,i,col1[i])
        #for i in range(len(comments)):
            #sheet1.write(i+1,0,ranks[i])
            #sheet1.write(i+1,1,dates[i])
            #sheet1.write(i+1,2,votes[i])
            #sheet1.write(i+1,3,comments[i])
        #sheet2=book.add_sheet('sheet2',cell_overwrite_ok=True)
        #col2=('电影名','评分','编号','评论数')
        #for i in range(0,4):
            #sheet2.write(0,i,col2[i])
        #sheet2.write(1,0,name)
        #sheet2.write(1,1,score)
        #sheet2.write(1,2,code)
        #sheet2.write(1,3,comment_num)
        #book.save(code + '.csv')
        #time.sleep(5)
'''
双线程运行
'''
t1 = threading.Thread(target = commentcaptcha, args= (0,8, 'driver'))
t2 = threading.Thread(target = commentcaptcha, args= (3,4, 'driver1' ))
'''
下为弃用线程，线程过多会导致豆瓣判断为机器人，封IP
t3 = threading.Thread(target = commentcaptcha, args= (40,60, 'driver2' ))
t4 = threading.Thread(target = commentcaptcha, args= (60,80, 'driver3' ))
t5 = threading.Thread(target = commentcaptcha, args= (80,100, 'driver4' ))
t6 = threading.Thread(target = commentcaptcha, args= (100,120, 'driver5' ))
'''
t1.start()
t2.start()



'''
电影信息补全
'''
from bs4 import BeautifulSoup as Soup
for path in pathes_new2:
    driver.get(path)
    source = driver.page_source
    soup = Soup(source,'lxml')
    
    director = soup.find_all('span',class_='attrs')[0]
    director = str(director).split('>')[-3].split('<')[0]           #导演
    try:
        editor = soup.find_all('span',class_='attrs')[1] 
        editor = re.findall('>[^(< )].+?<',str(editor))
        editors = ''
        for edit in editor:
            editor_name = edit.split('>')[-1].split('<')[0]
            editors += editor_name + r'/'                               #编剧
    except:editors = 'None'
    
    try:
        performer = soup.find_all('span',class_='attrs')[2]
        performer = re.findall('>[^(< )].+?<',str(performer))
        performers = ''
        for perform in performer:
            performer_name = perform.split('>')[-1].split('<')[0]
            performers += performer_name + r'/'                         #演员
    except:performers = 'None'
    
    classes = soup.find_all('span',property='v:genre')              #电影类型
    types = ''
    for typename in classes:
        typename = str(typename).split('>')[-2].split('<')[0]
        types += typename + r'/'
    
    filmname = soup.find_all('span',property='v:itemreviewed')
    filmname = str(filmname).split('>')[-2].split('<')[0]            #电影名
    
    date = soup.find_all('span',property="v:initialReleaseDate") 
    date = str(date).split('>')[-2].split('<')[0]                    #上映日期
       
    countries = re.findall(u'制片国家/地区.+',source)
    countries = countries[0].split('> ')[1].split('<')[0]            #制片国家
    
    film_info = [filmname,director,editors,performers,types,countries,date]
    
    book=Workbook()
    sheet1=book.active
    sheet1.title = "电影信息"
    col1=('电影名称','导演','编剧','主演','类型','制片国家/地区','上映时间')
    for i in range(1,8):
        sheet1.cell(row =1 ,column = i , value = col1[i-1])
        sheet1.cell(row =2 ,column = i , value = film_info[i-1])
    book.save(filmname + '.xlsx')
    time.sleep(1)
    





